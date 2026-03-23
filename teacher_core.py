# teacher_core.py  –  logic dùng chung cho cả local (tkinter) và Streamlit
import re
import json
import io
from datetime import datetime, timedelta, date as date_type
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===================== MÃ MÔN HỌC =====================
SUBJECT_MAP = {
    "ngữ văn": "NGUVAN", "van": "NGUVAN", "nguvan": "NGUVAN", "nv": "NGUVAN",
    "toán": "TOAN", "toan": "TOAN", "toán học": "TOAN",
    "tiếng anh": "ANH", "ngoại ngữ 1": "ANH", "ngoại ngữ 2": "ANH",
    "ngoại ngữ": "ANH", "anh": "ANH", "nn1": "ANH", "nn2": "ANH",
    "lịch sử": "LICHSU", "sử": "LICHSU", "lichsu": "LICHSU",
    "giáo dục thể chất": "GDTC", "thể dục": "GDTC", "gdtc": "GDTC", "td": "GDTC",
    "giáo dục quốc phòng và an ninh": "GDQP", "giáo dục quốc phòng": "GDQP",
    "qpan": "GDQP", "gdqp": "GDQP", "quốc phòng": "GDQP",
    "địa lí": "DIALY", "địa lý": "DIALY", "địa": "DIALY", "dialy": "DIALY",
    "giáo dục kinh tế và pháp luật": "GDKTPL", "gdktpl": "GDKTPL",
    "kinh tế pháp luật": "GDKTPL", "ktpl": "GDKTPL",
    "vật lí": "VATLY", "vật lý": "VATLY", "lí": "VATLY", "lý": "VATLY",
    "vatly": "VATLY", "vl": "VATLY",
    "hóa học": "HOAHOC", "hoá học": "HOAHOC", "hóa": "HOAHOC",
    "hoá": "HOAHOC", "hoahoc": "HOAHOC", "hh": "HOAHOC",
    "sinh học": "SINH", "sinh": "SINH",
    "cnnn": "CONGNGHE(NN)", "nông nghiệp": "CONGNGHE(NN)",
    "công nghệ (nn)": "CONGNGHE(NN)", "công nghệ(nn)": "CONGNGHE(NN)",
    "cncn": "CONGNGHE(CN)", "công nghiệp": "CONGNGHE(CN)",
    "công nghệ (cn)": "CONGNGHE(CN)", "công nghệ(cn)": "CONGNGHE(CN)",
    "công nghệ": "CONGNGHE", "cong nghe": "CONGNGHE",
    "tin học": "TINHOC", "tin": "TINHOC", "tinhoc": "TINHOC",
    "nội dung giáo dục địa phương": "NDGDDP", "giáo dục địa phương": "NDGDDP",
    "gdđp": "NDGDDP", "gddp": "NDGDDP",
    "hoạt động trải nghiệm, hướng nghiệp": "TNHN",
    "hoạt động trải nghiệm": "TNHN", "hđ trải nghiệm": "TNHN",
    "hđtn": "TNHN", "hđtn hn": "TNHN", "hdtn": "TNHN", "tnhn": "TNHN",
    "tiếng pháp": "TIENGPHAP", "tiếng nga": "TIENGNGA",
    "tiếng nhật": "TIENGNHAT", "tiếng trung": "TIENGTRUNG", "tiếng hàn": "TIENGHAN",
    "nghề phổ thông": "NGHEPHOTHONG", "nghề": "NGHEPHOTHONG",
    "âm nhạc": "AMNHAC", "nhạc": "AMNHAC",
    "mỹ thuật": "MYTHUAT", "mĩ thuật": "MYTHUAT", "mt": "MYTHUAT",
    "lịch sử và địa lí": "LICHSUDIALI", "lịch sử và địa lý": "LICHSUDIALI",
    "ls&đl": "LICHSUDIALI", "ls & đl": "LICHSUDIALI",
    "khoa học tự nhiên": "KHTN", "khtn": "KHTN",
    "giáo dục công dân": "GDCD", "gdcd": "GDCD",
    "hoạt động ngoài giờ lên lớp": "HDNGLL", "hđngll": "HDNGLL",
    "tiếng dân tộc thiểu số": "TDTTS",
    "nghệ thuật": "NGHETHUAT",
    "gd11dp": "NDGDDP",
}

_CLASS_PAT = r'\d{2}[A-Za-z]+\.?\d{0,2}(?:\.\d+)?'
_SUBJECT_STOPWORDS = {"đến", "den", "và", "từ", "lớp", "khối", "tới", "to", "the", "from"}

# ===================== SUBJECT LOOKUP =====================

def _normalize(text):
    return text.lower().strip() if text else ""

def match_subject_local(raw_subject):
    s = _normalize(raw_subject)
    if s in SUBJECT_MAP:
        return SUBJECT_MAP[s]
    best, best_len = None, 0
    for key, code in SUBJECT_MAP.items():
        if key in s or s in key:
            if len(key) > best_len:
                best, best_len = code, len(key)
    return best

def match_subject_ai(client, raw_subject):
    all_codes = list(set(SUBJECT_MAP.values()))
    prompt = (
        f'Bạn là chuyên gia nhận diện môn học trong hệ thống giáo dục Việt Nam.\n'
        f'Cho tên môn học thô: "{raw_subject}"\n'
        f'Hãy xác định đây là môn học nào trong danh sách mã môn sau:\n'
        f'{json.dumps(all_codes, ensure_ascii=False)}\n\n'
        f'Trả lời CHỈ bằng mã môn (ví dụ: TOAN, ANH, VATLY...) hoặc "UNKNOWN" nếu không xác định được.\n'
        f'Không giải thích, không thêm ký tự nào khác.'
    )
    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=50,
            messages=[{"role": "user", "content": prompt}]
        )
        result = response.content[0].text.strip().upper()
        if result in all_codes:
            return result
    except Exception:
        pass
    return None

def get_subject_code(client, raw_subject, ai_cache):
    if not raw_subject or not raw_subject.strip():
        return None
    raw_subject = raw_subject.strip()
    code = match_subject_local(raw_subject)
    if code:
        return code
    if raw_subject in ai_cache:
        return ai_cache[raw_subject]
    code = match_subject_ai(client, raw_subject)
    ai_cache[raw_subject] = code
    return code

# ===================== CLASS PARSING =====================

def expand_class_range(text):
    text = re.sub(r'(\d{2}[A-Za-z]+\d*)\(\d+\)', r'\1', text)
    classes = []

    range_pat = re.compile(
        r'(\d{2})([A-Za-zÀ-ỹ]+)(\d+)\s*(?:đến|den|-)\s*\1\2(\d+)', re.UNICODE
    )
    for m in range_pat.finditer(text):
        grade, alpha, start, end = m.groups()
        for i in range(int(start), int(end) + 1):
            classes.append(f"{grade}{alpha}{i}")
    text = range_pat.sub('', text)

    def _expand_compact(m):
        grade, alpha, digits = m.group(1), m.group(2), m.group(3)
        for d in digits:
            cls = f"{grade}{alpha}{d}"
            if cls not in classes:
                classes.append(cls)
        return ''

    text = re.sub(r'(\d{2})([A-Za-z]+)(\d{3,})', _expand_compact, text)
    text = re.sub(r'(\d{2})([A-Za-z]+)(\d{2})(?![,;.\s])', _expand_compact, text)

    def _expand_suffix(m):
        base, nums = m.group(1), m.group(2)
        for n in re.split(r'[,\s]+', nums):
            if n:
                classes.append(f"{base}{n.strip()}")
        return ''

    text = re.sub(r'(\d{2}[A-Za-z]+)(\d(?:,\s*\d)+)(?!\d)', _expand_suffix, text)
    classes.extend(re.findall(_CLASS_PAT, text))

    result, seen = [], set()
    for c in classes:
        c = c.strip().strip(',').strip()
        if c and c not in seen:
            seen.add(c)
            result.append(c)
    return result


def parse_pccm(raw_pccm):
    if not raw_pccm or (isinstance(raw_pccm, float) and pd.isna(raw_pccm)):
        return []

    text = str(raw_pccm).strip()

    def expand_parens(m):
        inner = m.group(1).strip()
        if re.fullmatch(r'\d+', inner):
            return ''
        return ',' + inner + ','

    text = re.sub(r'\(([^)]*)\)', expand_parens, text)
    text = text.replace(';', ',').replace('\n', '+')

    CLASS_RANGE_PAT = (
        r'\d{2}[A-Za-z]+\d+\s*(?:đến|den|-)\s*\d{2}[A-Za-z]+\d+'
        r'|\d{2}[A-Za-z]+\d{3,}'
        r'|' + _CLASS_PAT
    )
    tokens = []
    results = []
    token_re = re.compile(
        r'(?P<class>' + CLASS_RANGE_PAT + r')'
        r'|(?P<sep>[+,\s]+)'
        r'|(?P<colon>:)'
        r'|(?P<word>[A-Za-zÀ-ỹĐđ][A-Za-zÀ-ỹĐđ\(\)]*)'
        r'|(?P<other>.)',
        re.UNICODE
    )
    for m in token_re.finditer(text):
        tokens.append((m.lastgroup, m.group().strip()))

    merged = []
    i = 0
    while i < len(tokens):
        kind, val = tokens[i]
        if kind == 'word':
            words = [val]
            j = i + 1
            while j < len(tokens):
                k2, v2 = tokens[j]
                if k2 == 'word':
                    words.append(v2); j += 1
                elif k2 == 'sep' and j + 1 < len(tokens) and tokens[j+1][0] == 'word':
                    words.append(tokens[j+1][1]); j += 2
                else:
                    break
            merged.append(('word', ' '.join(words)))
            i = j
        elif kind == 'sep':
            if val:
                merged.append(('sep', val))
            i += 1
        elif kind in ('class', 'colon', 'other'):
            merged.append((kind, val))
            i += 1
        else:
            i += 1

    current_subject = None
    current_classes = []

    def flush(subj, classes, out):
        if subj and classes:
            out.append((subj, classes))
        elif classes:
            out.append(("", classes))

    idx = 0
    while idx < len(merged):
        kind, val = merged[idx]
        if kind == 'word':
            if val.lower().strip() in _SUBJECT_STOPWORDS:
                idx += 1
                continue
            next_non_sep = None
            for k2, v2 in merged[idx+1:]:
                if k2 != 'sep':
                    next_non_sep = (k2, v2)
                    break
            if next_non_sep and next_non_sep[0] == 'colon':
                flush(current_subject, current_classes, results)
                current_subject = val
                current_classes = []
                idx += 1
                while idx < len(merged) and merged[idx][0] in ('sep', 'colon'):
                    idx += 1
            elif next_non_sep and next_non_sep[0] == 'class':
                flush(current_subject, current_classes, results)
                current_subject = val
                current_classes = []
                idx += 1
                while idx < len(merged) and merged[idx][0] == 'sep':
                    idx += 1
            else:
                idx += 1
        elif kind == 'class':
            current_classes.extend(expand_class_range(val))
            idx += 1
        elif kind in ('sep', 'colon', 'other'):
            idx += 1
        else:
            idx += 1

    flush(current_subject, current_classes, results)
    return results

# ===================== UTILITIES =====================

def format_date(val):
    try:
        if val is None:
            return None, ""
        if isinstance(val, datetime):
            return val, val.strftime("%d/%m/%Y")
        if isinstance(val, date_type):
            dt = datetime(val.year, val.month, val.day)
            return dt, dt.strftime("%d/%m/%Y")
        if isinstance(val, (int, float)):
            if pd.isna(val):
                return None, ""
            dt = datetime(1899, 12, 30) + timedelta(days=int(val))
            return dt, dt.strftime("%d/%m/%Y")
        s = str(val).strip()
        if not s or s.lower() in ("nan", "nat", "none", ""):
            return None, ""
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt, dt.strftime("%d/%m/%Y")
            except Exception:
                pass
        return None, s
    except Exception:
        return None, ""


def find_column(df, candidates):
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        c = cand.lower().strip()
        if c in cols_lower:
            return cols_lower[c]
        for key, orig in cols_lower.items():
            if c in key or key in c:
                return orig
    return None


def detect_header_row(sheet_df):
    header_keywords = ['stt', 'họ tên', 'họ và tên', 'giáo viên', 'pccm', 'phân công', 'ngày sinh']
    for i, row in sheet_df.iterrows():
        vals = [str(v).lower().strip() for v in row.values if pd.notna(v)]
        matches = sum(1 for v in vals for kw in header_keywords if kw in v)
        if matches >= 2:
            return i
    return 0


def get_grade(class_name):
    """Lấy khối từ tên lớp: '10A1' -> 10, '11C' -> 11"""
    m = re.match(r'^(\d{2})', str(class_name).strip())
    return int(m.group(1)) if m else None

# ===================== STYLING =====================

def _style_header(ws, row, ncols, color="1F4E79"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

def _style_data_row(ws, row, ncols, even, left_cols=()):
    fill = PatternFill("solid", fgColor="EBF3FB" if even else "FFFFFF")
    font = Font(name="Arial", size=10)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = (
            Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col in left_cols
            else Alignment(horizontal="center", vertical="center")
        )

def _add_borders(ws, start_row, end_row, ncols):
    thin = Side(style='thin', color='B0C4DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).border = border

# ===================== MAIN PROCESSING =====================

def process_data(input_bytes_or_path, nien_khoa, client, progress_cb=None):
    """
    Xử lý file Excel đầu vào.
    input_bytes_or_path: bytes (Streamlit) hoặc str path (local)
    nien_khoa: chuỗi vd "2025-2026"
    client: anthropic.Anthropic()
    progress_cb: callback(msg: str) để report tiến độ (optional)
    Trả về: bytes của file Excel đầu ra
    """
    def log(msg):
        if progress_cb:
            progress_cb(msg)

    # Đọc file
    if isinstance(input_bytes_or_path, (bytes, bytearray, io.BytesIO)):
        src = io.BytesIO(input_bytes_or_path) if isinstance(input_bytes_or_path, (bytes, bytearray)) else input_bytes_or_path
    else:
        src = input_bytes_or_path

    xl = pd.ExcelFile(src)
    sheet_names = xl.sheet_names
    data_sheet = next((s for s in sheet_names if s.strip().lower() == "data"), sheet_names[0])

    log(f"Đọc sheet '{data_sheet}'...")
    raw_df = pd.read_excel(src, sheet_name=data_sheet, header=None)
    header_row_idx = detect_header_row(raw_df)
    df = pd.read_excel(src, sheet_name=data_sheet, header=header_row_idx)
    df.columns = [str(c).strip() for c in df.columns]

    col_stt    = find_column(df, ["stt", "tt", "số thứ tự", "no"])
    col_hoten  = find_column(df, ["họ tên", "họ và tên", "tên", "giáo viên", "ho ten", "hoten"])
    col_ngay   = find_column(df, ["ngày sinh", "ngay sinh", "sinh ngày", "dob", "birthday"])
    col_pccm   = find_column(df, ["pccm", "phân công chuyên môn", "phân công",
                                   "giảng dạy lớp", "môn học giảng dạy", "phan cong", "giang day"])

    if not col_hoten:
        raise ValueError("Không tìm thấy cột Họ tên trong sheet Data!")
    if not col_pccm:
        raise ValueError("Không tìm thấy cột PCCM trong sheet Data!")

    df = df[df[col_hoten].notna() & (df[col_hoten].astype(str).str.strip() != "")].copy()
    df = df.reset_index(drop=True)

    ai_cache = {}
    teachers = []
    total = len(df)

    for idx, row in df.iterrows():
        log(f"Xử lý giáo viên {idx+1}/{total}: {row[col_hoten]}")
        stt = str(row[col_stt]).strip() if col_stt and pd.notna(row.get(col_stt)) else str(idx + 1)
        ho_ten = str(row[col_hoten]).strip()
        ngay_dt, ngay_str = format_date(row[col_ngay]) if col_ngay and pd.notna(row.get(col_ngay)) else (None, "")
        pccm_raw = str(row[col_pccm]).strip() if pd.notna(row.get(col_pccm)) else ""

        parsed = parse_pccm(pccm_raw)
        subject_codes = []
        mon_lop_list = []

        for subj_raw, lop_list in parsed:
            code = get_subject_code(client, subj_raw, ai_cache)
            if code:
                if code not in subject_codes:
                    subject_codes.append(code)
                for lop in lop_list:
                    lop = lop.strip()
                    if lop:
                        mon_lop_list.append((lop, code))
            else:
                for lop in lop_list:
                    lop = lop.strip()
                    if lop:
                        mon_lop_list.append((lop, subj_raw.upper() if subj_raw else "?"))

        # Loại trùng trong cùng GV
        seen_pairs = set()
        unique_ml = []
        for lop, code in mon_lop_list:
            if (lop, code) not in seen_pairs:
                seen_pairs.add((lop, code))
                unique_ml.append((lop, code))

        teachers.append({
            "stt": stt,
            "ho_ten": ho_ten,
            "ngay_dt": ngay_dt,
            "ngay_str": ngay_str,
            "subject_codes": subject_codes,
            "mon_lop_list": unique_ml,
        })

    # Phát hiện trùng tổ hợp môn-lớp giữa các GV
    pair_count = defaultdict(list)
    for t in teachers:
        for lop, code in t["mon_lop_list"]:
            pair_count[(lop, code)].append(t["ho_ten"])

    for t in teachers:
        parts = []
        for lop, code in t["mon_lop_list"]:
            key = (lop, code)
            if len(pair_count[key]) > 1:
                parts.append(f"{lop}-{code}({t['ho_ten']})")
            else:
                parts.append(f"{lop}-{code}")
        t["pccm_str"] = ",".join(parts)

    # Tổng hợp tất cả lớp duy nhất (sắp xếp theo khối rồi tên lớp)
    all_classes_set = set()
    for t in teachers:
        for lop, _ in t["mon_lop_list"]:
            all_classes_set.add(lop.strip())
    all_classes = sorted(all_classes_set, key=lambda c: (get_grade(c) or 99, c))

    log("Tạo file Excel đầu ra...")
    wb = openpyxl.Workbook()

    # ── Sheet CLASS (bên trái Teachers) ──────────────────────────────
    ws_class = wb.active
    ws_class.title = "Class"

    # Hàng 1: Niên khóa
    ws_class["A1"] = "Niên khóa"
    ws_class["B1"] = nien_khoa
    # Hàng 2: Lớp / Khối header
    ws_class["A2"] = "Lớp"
    ws_class["B2"] = "Khối"

    # Style header row 1-2
    for r in (1, 2):
        for col in ("A", "B"):
            cell = ws_class[f"{col}{r}"]
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dữ liệu lớp từ A3
    for i, cls in enumerate(all_classes):
        r = i + 3
        ws_class.cell(row=r, column=1, value=cls)
        ws_class.cell(row=r, column=2, value=get_grade(cls))
        for col in (1, 2):
            cell = ws_class.cell(row=r, column=col)
            cell.fill = PatternFill("solid", fgColor="EBF3FB" if i % 2 == 0 else "FFFFFF")
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    _add_borders(ws_class, 1, len(all_classes) + 2, 2)
    ws_class.column_dimensions["A"].width = 14
    ws_class.column_dimensions["B"].width = 10
    ws_class.freeze_panes = "A3"

    # ── Sheet TEACHERS ────────────────────────────────────────────────
    ws_t = wb.create_sheet("Teachers")

    # Cột: STT | Họ tên | Ngày sinh | SĐT | Môn dạy | TBM | CN | PCCM
    headers_t = ["STT", "Họ tên", "Ngày sinh", "SĐT", "Môn dạy", "TBM", "CN", "PCCM"]
    for ci, h in enumerate(headers_t, 1):
        ws_t.cell(row=1, column=ci, value=h)
    _style_header(ws_t, 1, len(headers_t))
    ws_t.row_dimensions[1].height = 30

    for i, t in enumerate(teachers):
        rn = i + 2
        ws_t.cell(row=rn, column=1, value=t["stt"])
        ws_t.cell(row=rn, column=2, value=t["ho_ten"])
        dc = ws_t.cell(row=rn, column=3)
        if t["ngay_dt"]:
            dc.value = t["ngay_dt"]
            dc.number_format = "DD/MM/YYYY"
        else:
            dc.value = t["ngay_str"]
        ws_t.cell(row=rn, column=4, value="")   # SĐT
        ws_t.cell(row=rn, column=5, value=", ".join(t["subject_codes"]))  # Môn dạy
        ws_t.cell(row=rn, column=6, value="")   # TBM
        ws_t.cell(row=rn, column=7, value="")   # CN
        ws_t.cell(row=rn, column=8, value=t["pccm_str"])
        _style_data_row(ws_t, rn, len(headers_t), i % 2 == 0, left_cols=(2, 5, 8))

    _add_borders(ws_t, 1, len(teachers) + 1, len(headers_t))
    col_widths_t = [6, 25, 14, 14, 30, 10, 10, 80]
    for ci, w in enumerate(col_widths_t, 1):
        ws_t.column_dimensions[get_column_letter(ci)].width = w
    ws_t.freeze_panes = "A2"

    # ── Sheet STUDENTS (bên phải Teachers) ───────────────────────────
    ws_s = wb.create_sheet("Students")
    headers_s = ["STT", "Mã HS", "Họ tên", "Lớp", "Giới tính",
                 "Ngày sinh", "Số điện thoại", "Email", "Tài khoản"]
    for ci, h in enumerate(headers_s, 1):
        ws_s.cell(row=1, column=ci, value=h)
    _style_header(ws_s, 1, len(headers_s))
    ws_s.row_dimensions[1].height = 30
    _add_borders(ws_s, 1, 1, len(headers_s))
    col_widths_s = [6, 14, 25, 10, 12, 14, 16, 28, 18]
    for ci, w in enumerate(col_widths_s, 1):
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    ws_s.freeze_panes = "A2"

    # Đặt thứ tự sheet: Class | Teachers | Students
    # wb.worksheets thứ tự: ws_class(0), ws_t(1), ws_s(2) — đúng rồi

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    log("Hoàn thành!")
    return out.read()
